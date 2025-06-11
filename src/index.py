
import boto3
import json
import re
from datetime import datetime, timedelta
from datetime import datetime, date
from io import BytesIO
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from botocore.exceptions import ClientError
import base64
import traceback
import os
import uuid

# Environment variables for configuration
BEDROCK_MODEL_ID = os.environ.get('BEDROCK_MODEL_ID', 'anthropic.claude-3-sonnet-20240229-v1:0')
BEDROCK_TEMPERATURE = float(os.environ.get('BEDROCK_TEMPERATURE', '0.1'))
BEDROCK_TOP_P = float(os.environ.get('BEDROCK_TOP_P', '0.9'))
BEDROCK_MAX_TOKENS = int(os.environ.get('BEDROCK_MAX_TOKENS', '4000'))
EXCEL_FILENAME_TEMPLATE = os.environ.get('EXCEL_FILENAME_TEMPLATE', 'AWS_Health_Events_Analysis_{date}_{time}.xlsx')
customer_name = os.environ.get('CUSTOMER_NAME', 'Notification')
excluded_services_str = os.environ.get('EXCLUDED_SERVICES', '')
S3_BUCKET_NAME = os.environ.get('S3_BUCKET_NAME', '')
S3_KEY_PREFIX = os.environ.get('S3_KEY_PREFIX', '')

excluded_services = [s.strip() for s in excluded_services_str.split(',') if s.strip()]


def expand_events_by_account(events):
    """
    Expands events that affect multiple accounts into separate event records for each account.
    Fetches affected accounts if not already specified.
    
    Args:
        events (list): List of event dictionaries
        
    Returns:
        list: Expanded list of event dictionaries
    """
    expanded_events = []
    health_client = boto3.client('health', region_name='us-east-1')
    
    for event in events:
        # Get the account ID string which may contain multiple comma-separated IDs
        account_id_str = event.get('accountId', '')
        event_arn = event.get('arn', '')
        
        # If no account ID or it's N/A, try to fetch affected accounts
        if not account_id_str or account_id_str == 'N/A':
            try:
                print(f"Fetching affected accounts for event: {event.get('eventTypeCode', 'unknown')}")
                response = health_client.describe_affected_accounts_for_organization(
                    eventArn=event_arn
                )
                affected_accounts = response.get('affectedAccounts', [])
                
                if affected_accounts:
                    # If multiple accounts are affected, join them with commas
                    account_id_str = ', '.join(affected_accounts)
                    event['accountId'] = account_id_str  # Update the event with the account IDs
                    print(f"Found affected accounts: {account_id_str}")
                else:
                    print("No affected accounts found")
                    expanded_events.append(event)  # Keep the event as is
                    continue
            except Exception as e:
                print(f"Error fetching affected accounts: {str(e)}")
                expanded_events.append(event)  # Keep the event as is
                continue
        
        # If no comma in the string, it's a single account or none
        if ',' not in account_id_str:
            expanded_events.append(event)
            continue
        
        # Split the account IDs and create a separate event for each
        account_ids = [aid.strip() for aid in account_id_str.split(',')]
        print(f"Expanding event {event_arn} for {len(account_ids)} accounts: {account_ids}")
        
        for account_id in account_ids:
            # Create a copy of the event for this specific account
            account_event = event.copy()
            account_event['accountId'] = account_id
            expanded_events.append(account_event)
    
    print(f"Expanded {len(events)} events to {len(expanded_events)} account-specific events")
    return expanded_events

def lambda_handler(event, context):
    print("Starting execution...")
    
    try:
        # Get all configuration from environment variables
        analysis_window_days = int(os.environ['ANALYSIS_WINDOW_DAYS'])
        
        # Get event categories to process from environment variable
        event_categories_to_process = []
        if 'EVENT_CATEGORIES' in os.environ and os.environ['EVENT_CATEGORIES'].strip():
            event_categories_to_process = [cat.strip() for cat in os.environ['EVENT_CATEGORIES'].split(',')]
            print(f"Will only process these event categories: {event_categories_to_process}")
        else:
            print("No EVENT_CATEGORIES specified, will process all event categories")

        excluded_services_str = os.environ.get('EXCLUDED_SERVICES', '')
        excluded_services = [s.strip() for s in excluded_services_str.split(',') if s.strip()]
        
        if excluded_services:
            print(f"Excluding services from analysis: {excluded_services}")
        
        # Set up time range for filtering using environment variable
        bedrock_client = get_bedrock_client()
        end_time = datetime.utcnow()
        start_time = end_time - timedelta(days=analysis_window_days)
        
        print(f"Fetching events between {start_time} and {end_time}")
        
        # Format dates properly for the API
        formatted_start = start_time.strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + 'Z'
        formatted_end = end_time.strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + 'Z'
        
        # Initialize AWS Health client
        health_client = boto3.client('health', region_name='us-east-1')
        
        # Initialize variables for event collection
        all_events = []
        filtered_count = 0
        
        try:
            # Check if we should use organization view or account view
            use_org_view = is_org_view_enabled()
            
            if use_org_view:
                print("Using AWS Health Organization View")
                
                # CHANGE 1: Fetch closed events with both start and end date filters
                closed_filter = {
                    'startTime': {'from': formatted_start},
                    'endTime': {'to': formatted_end},
                    'eventStatusCodes': ['closed', 'upcoming']
                }
                
                # Add event type categories filter if specified
                if event_categories_to_process:
                    closed_filter['eventTypeCategories'] = event_categories_to_process
                
                print(f"Fetching CLOSED events with filter: {closed_filter}")
                closed_response = health_client.describe_events_for_organization(
                    filter=closed_filter,
                    maxResults=100
                )
                
                if 'events' in closed_response:
                    all_events.extend(closed_response['events'])
                    print(f"Retrieved {len(closed_response.get('events', []))} closed events")
                
                # Handle pagination for closed events
                while 'nextToken' in closed_response and closed_response['nextToken']:
                    print(f"Found nextToken for closed events, fetching more...")
                    if context.get_remaining_time_in_millis() < 15000:  # 15 seconds buffer
                        print("Approaching Lambda timeout, stopping pagination")
                        break
                        
                    closed_response = health_client.describe_events_for_organization(
                        filter=closed_filter,
                        maxResults=100,
                        nextToken=closed_response['nextToken']
                    )
                    
                    if 'events' in closed_response:
                        all_events.extend(closed_response['events'])
                        print(f"Retrieved {len(closed_response.get('events', []))} additional closed events")
                
                # CHANGE 2: Fetch open events with only start date filter
                open_filter = {
                    'startTime': {'from': formatted_start},  # Started on or after start date
                    'eventStatusCodes': ['open']  # Only open events
                }
                
                # Add event type categories filter if specified
                if event_categories_to_process:
                    open_filter['eventTypeCategories'] = event_categories_to_process
                
                print(f"Fetching OPEN events with filter: {open_filter}")
                open_response = health_client.describe_events_for_organization(
                    filter=open_filter,
                    maxResults=100
                )
                
                if 'events' in open_response:
                    all_events.extend(open_response['events'])
                    print(f"Retrieved {len(open_response.get('events', []))} open events")
                
                # Handle pagination for open events
                while 'nextToken' in open_response and open_response['nextToken']:
                    print(f"Found nextToken for open events, fetching more...")
                    if context.get_remaining_time_in_millis() < 15000:  # 15 seconds buffer
                        print("Approaching Lambda timeout, stopping pagination")
                        break
                        
                    open_response = health_client.describe_events_for_organization(
                        filter=open_filter,
                        maxResults=100,
                        nextToken=open_response['nextToken']
                    )
                    
                    if 'events' in open_response:
                        all_events.extend(open_response['events'])
                        print(f"Retrieved {len(open_response.get('events', []))} additional open events")
                
            else:
                print("Using AWS Health Account View")
                
                # CHANGE 3: Same approach for account view - fetch closed events
                closed_filter = {
                    'startTime': {'from': formatted_start},
                    'endTime': {'to': formatted_end},
                    'eventStatusCodes': ['closed', 'upcoming']
                }
                
                # Add event type categories filter if specified
                if event_categories_to_process:
                    closed_filter['eventTypeCategories'] = event_categories_to_process
                
                print(f"Fetching CLOSED events with filter: {closed_filter}")
                closed_response = health_client.describe_events(
                    filter=closed_filter,
                    maxResults=100
                )
                
                if 'events' in closed_response:
                    all_events.extend(closed_response['events'])
                
                # Handle pagination for closed events
                while 'nextToken' in closed_response and closed_response['nextToken']:
                    if context.get_remaining_time_in_millis() < 15000:  # 15 seconds buffer
                        print("Approaching Lambda timeout, stopping pagination")
                        break
                        
                    closed_response = health_client.describe_events(
                        filter=closed_filter,
                        maxResults=100,
                        nextToken=closed_response['nextToken']
                    )
                    
                    if 'events' in closed_response:
                        all_events.extend(closed_response['events'])
                
                # CHANGE 4: Fetch open events with only start date filter
                open_filter = {
                    'startTime': {'from': formatted_start},  # Started on or after start date
                    'eventStatusCodes': ['open']  # Only open events
                }
                
                # Add event type categories filter if specified
                if event_categories_to_process:
                    open_filter['eventTypeCategories'] = event_categories_to_process
                
                print(f"Fetching OPEN events with filter: {open_filter}")
                open_response = health_client.describe_events(
                    filter=open_filter,
                    maxResults=100
                )
                
                if 'events' in open_response:
                    all_events.extend(open_response['events'])
                
                # Handle pagination for open events
                while 'nextToken' in open_response and open_response['nextToken']:
                    if context.get_remaining_time_in_millis() < 15000:  # 15 seconds buffer
                        print("Approaching Lambda timeout, stopping pagination")
                        break
                        
                    open_response = health_client.describe_events(
                        filter=open_filter,
                        maxResults=100,
                        nextToken=open_response['nextToken']
                    )
                    
                    if 'events' in open_response:
                        all_events.extend(open_response['events'])
        
        except ClientError as e:
            if e.response['Error']['Code'] == 'SubscriptionRequiredException':
                print("Health Organization View is not enabled. Falling back to account-specific view.")
                
                # CHANGE 5: Same approach for fallback - fetch closed events
                closed_filter = {
                    'startTime': {'from': formatted_start},
                    'endTime': {'to': formatted_end},
                    'eventStatusCodes': ['closed', 'upcoming']
                }
                
                # Add event type categories filter if specified
                if event_categories_to_process:
                    closed_filter['eventTypeCategories'] = event_categories_to_process
                
                print(f"Fetching CLOSED events with filter: {closed_filter}")
                closed_response = health_client.describe_events(
                    filter=closed_filter,
                    maxResults=100
                )
                
                if 'events' in closed_response:
                    all_events.extend(closed_response['events'])
                
                # Handle pagination for closed events
                while 'nextToken' in closed_response and closed_response['nextToken']:
                    if context.get_remaining_time_in_millis() < 15000:  # 15 seconds buffer
                        print("Approaching Lambda timeout, stopping pagination")
                        break
                        
                    closed_response = health_client.describe_events(
                        filter=closed_filter,
                        maxResults=100,
                        nextToken=closed_response['nextToken']
                    )
                    
                    if 'events' in closed_response:
                        all_events.extend(closed_response['events'])
                
                # CHANGE 6: Fetch open events with only start date filter
                open_filter = {
                    'startTime': {'from': formatted_start},  # Started on or after start date
                    'eventStatusCodes': ['open']  # Only open events
                }
                
                # Add event type categories filter if specified
                if event_categories_to_process:
                    open_filter['eventTypeCategories'] = event_categories_to_process
                
                print(f"Fetching OPEN events with filter: {open_filter}")
                open_response = health_client.describe_events(
                    filter=open_filter,
                    maxResults=100
                )
                
                if 'events' in open_response:
                    all_events.extend(open_response['events'])
                
                # Handle pagination for open events
                while 'nextToken' in open_response and open_response['nextToken']:
                    if context.get_remaining_time_in_millis() < 15000:  # 15 seconds buffer
                        print("Approaching Lambda timeout, stopping pagination")
                        break
                        
                    open_response = health_client.describe_events(
                        filter=open_filter,
                        maxResults=100,
                        nextToken=open_response['nextToken']
                    )
                    
                    if 'events' in open_response:
                        all_events.extend(open_response['events'])
            else:
                raise
        
        # CHANGE 7: Remove duplicates by ARN
        unique_events = {}
        for item in all_events:
            arn = item.get('arn')
            if arn and arn not in unique_events:
                unique_events[arn] = item
        
        all_events = list(unique_events.values())

        # Filter out excluded services post-retrieval
        if excluded_services:
            filtered_events = [e for e in all_events if e.get('service') not in excluded_services]
            print(f"Filtered out {len(all_events) - len(filtered_events)} events from excluded services")
            all_events = filtered_events
        
        # NEW STEP: Expand events for multiple accounts
        all_events_original = all_events.copy()
        all_events_expanded = expand_events_by_account(all_events)
        print(f"Expanded {len(all_events_original)} events to {len(all_events_expanded)} account-specific events")
        
        items_count = len(all_events_original)  # Keep the original count for reporting
        print(f"Fetched {items_count} unique events from AWS Health API (expanded to {len(all_events_expanded)} account-specific events)")
        
        if len(all_events_expanded) == 0:
            print("No events found with filter")
            return {
                'statusCode': 200,
                'body': json.dumps({
                    'message': 'No events matched the filter criteria',
                    'filters_used': {
                        'closed_filter': closed_filter if 'closed_filter' in locals() else {},
                        'open_filter': open_filter if 'open_filter' in locals() else {}
                    }
                })
            }
        
        # Process events directly from the API results
        events_analysis = []
        event_categories = defaultdict(int)
        raw_events = []  # Store raw event data for Excel
        
        # Process each event from the expanded API results
        for item in all_events_expanded:
            if context.get_remaining_time_in_millis() > 10000:
                # Check if we should process this event category
                event_type_category = item.get('eventTypeCategory', '')
                
                # Skip events that don't match our configured categories (this is redundant since we're filtering in the API call,
                # but keeping it for consistency with the original code)
                if event_categories_to_process and event_type_category not in event_categories_to_process:
                    print(f"Skipping event {item.get('eventTypeCode', 'unknown')} with category {event_type_category} (not in configured categories)")
                    filtered_count += 1
                    continue
                
                print(f"Processing event: {item.get('eventTypeCode', 'unknown')} with category {event_type_category}")
                
                try:
                    # Store raw event data for Excel
                    raw_events.append(item)
                    
                    # Ensure we have the event ARN and standardize field name
                    event_arn = item.get('arn', '')
                    if event_arn:
                        item['eventArn'] = event_arn  # Standardize field name
                
                    # Extract account ID from ARN - this is already handled by the expansion function
                    account_id = item.get('accountId', 'N/A')
                    print(f"Processing with account ID: {account_id}")
                    
                    # Fetch additional details from Health API - now with a single account ID
                    health_data = fetch_health_event_details1(item.get('arn', ''), account_id)
                    
                    # Extract the actual description for analysis - IMPROVED EXTRACTION
                    actual_description = health_data['details'].get('eventDescription', {}).get('latestDescription', '')
                    
                    # If no description from Health API, try other possible fields
                    if not actual_description:
                        actual_description = (
                            item.get('eventDescription', '') or 
                            item.get('description', '') or 
                            item.get('message', '') or
                            'No description available'
                        )
                    
                    # Log the description we found
                    print(f"Using description (length: {len(actual_description)}): {actual_description[:100]}...")
                    
                    # Update the item with the actual description to improve analysis
                    item_with_description = item.copy()
                    item_with_description['description'] = actual_description
                    
                    analysis = analyze_event_with_bedrock(bedrock_client, item_with_description)
                    
                    categories = categorize_analysis(analysis)
                    if categories.get('critical', False):
                        event_categories['critical'] += 1
                    
                    risk_level = categories.get('risk_level', 'low')
                    event_categories[f"{risk_level}_risk"] += 1
                    
                    account_impact = categories.get('account_impact', 'low')
                    event_categories[f"{account_impact}_impact"] += 1
                    
                    # Create structured event data with both raw data and analysis
                    event_entry = {
                        "arn": item.get('arn', 'N/A'),
                        "eventArn": item.get('eventArn', item.get('arn', 'N/A')),  # Ensure eventArn is included
                        "event_type": item.get('eventTypeCode', 'N/A'),
                        "description": actual_description,
                        "region": item.get('region', 'N/A'),
                        "start_time": format_time(item.get('startTime', 'N/A')),
                        "last_update_time": format_time(item.get('lastUpdatedTime', 'N/A')),
                        "event_type_category": item.get('eventTypeCategory', 'N/A'),
                        "analysis_text": analysis,
                        "critical": categories.get('critical', False),
                        "risk_level": categories.get('risk_level', 'low'),
                        "accountId": account_id,  # Use the single account ID from the expanded event
                        "impact_analysis": categories.get('impact_analysis', ''),
                        "required_actions": categories.get('required_actions', ''),
                        "time_sensitivity": categories.get('time_sensitivity', 'Routine'),
                        "risk_category": categories.get('risk_category', 'Unknown'),
                        "consequences_if_ignored": categories.get('consequences_if_ignored', ''),
                        "affected_resources": extract_affected_resources(health_data['entities'])
                    }
                    
                    events_analysis.append(event_entry)
                    print(f"Successfully analyzed event {len(events_analysis)}")
                except Exception as e:
                    print(f"Error analyzing event: {str(e)}")
                    traceback.print_exc()
            else:
                print("Approaching Lambda timeout, stopping event processing")
                break
        
        if events_analysis:
            print(f"Successfully analyzed {len(events_analysis)} events (filtered out {filtered_count} events)")
            
            # Create Excel report with structured data
            excel_buffer = create_excel_report_improved(events_analysis)
            
            # Generate summary HTML with filtering info
            summary_html = generate_summary_html(
                items_count,  # Use original count before expansion
                event_categories, 
                filtered_count, 
                event_categories_to_process if event_categories_to_process else None,
                events_analysis  
            )
            
            # Send email with attachment
            send_ses_email_with_attachment(summary_html, excel_buffer, items_count, event_categories, events_analysis)
            
            try:
                # Add CloudWatch metrics with error handling
                add_cloudwatch_metrics(event_categories, len(events_analysis), items_count, filtered_count)
            except Exception as e:
                print(f"Error publishing CloudWatch metrics: {str(e)}")
                print("Continuing execution despite metrics error")
            
            return {
                'statusCode': 200,
                'body': json.dumps({
                    'total_events': items_count,  # Original event count
                    'total_expanded_events': len(all_events_expanded),  # Expanded event count
                    'analyzed_events': len(events_analysis),
                    'filtered_events': filtered_count,
                    'categories': dict(event_categories),
                    'category_filter_applied': bool(event_categories_to_process),
                    'categories_processed': event_categories_to_process,
                    'filters_used': {
                        'closed_filter': closed_filter if 'closed_filter' in locals() else {},
                        'open_filter': open_filter if 'open_filter' in locals() else {}
                    }
                })
            }
        else:
            print(f"No events were successfully analyzed (filtered out {filtered_count} events)")
            return {
                'statusCode': 200,
                'body': json.dumps({
                    'message': 'Found events but none were analyzed',
                    'events_found': items_count,
                    'expanded_events_found': len(all_events_expanded),
                    'filtered_events': filtered_count,
                    'category_filter_applied': bool(event_categories_to_process),
                    'categories_processed': event_categories_to_process,
                    'filters_used': {
                        'closed_filter': closed_filter if 'closed_filter' in locals() else {},
                        'open_filter': open_filter if 'open_filter' in locals() else {}
                    }
                })
            }
            
    except Exception as e:
        print(f"Error: {str(e)}")
        traceback.print_exc()
        return {
            'statusCode': 500,
            'body': json.dumps({'error': str(e)})
        }



def is_org_view_enabled():
    """
    Check if AWS Health Organization View is enabled
    
    Returns:
        bool: True if organization view is enabled, False otherwise
    """
    try:
        # Try to call an organization-specific API to check if it's enabled
        health_client = boto3.client('health', region_name='us-east-1')
        # This will throw an exception if org view is not enabled
        health_client.describe_events_for_organization(
            filter={},
            maxResults=1
        )
        return True
    except Exception as e:
        error_code = getattr(e, 'response', {}).get('Error', {}).get('Code', '')
        if error_code == 'SubscriptionRequiredException':
            return False
        # For any other error, assume we don't have org view permissions
        return False

def get_bedrock_client():
    """
    Get Amazon Bedrock client
    
    Returns:
        boto3.client: Bedrock runtime client
    """
    return boto3.client(service_name='bedrock-runtime',region_name='us-east-1')

def format_time(time_str):
    """
    Format time string to be consistent
    
    Args:
        time_str (str): ISO format time string
        
    Returns:
        str: Formatted time string (YYYY-MM-DD)
    """
    if not time_str or time_str == 'N/A':
        return 'N/A'
    
    try:
        # If it's already a datetime object
        if isinstance(time_str, datetime):
            return time_str.strftime('%Y-%m-%d')
        
        # Parse ISO format
        dt = datetime.fromisoformat(time_str.replace('Z', '+00:00'))
        return dt.strftime('%Y-%m-%d')
    except Exception:
        # If we can't parse it, return as is
        return time_str

def fetch_health_event_details(event_arn):
    """
    Fetch detailed event information from AWS Health API
    
    Args:
        event_arn (str): ARN of the health event
        
    Returns:
        dict: Event details including affected resources
    """
    try:
        health_client = boto3.client('health', region_name='us-east-1')
        
        # Get event details
        event_details = health_client.describe_event_details(
            eventArns=[event_arn]
        )
        
        # Get affected entities
        affected_entities = health_client.describe_affected_entities(
            filter={
                'eventArns': [event_arn]
            }
        )
        
        return {
            'details': event_details.get('successfulSet', [{}])[0] if event_details.get('successfulSet') else {},
            'entities': affected_entities.get('entities', [])
        }
    except Exception as e:
        print(f"Error fetching Health API data: {str(e)}")
        return {'details': {}, 'entities': []}
         
def extract_affected_resources(entities):
    """
    Extract affected resources from Health API entities
    
    Args:
        entities (list): List of entity objects from Health API
        
    Returns:
        str: Comma-separated list of affected resources
    """
    if not entities:
        return "None specified"
    
    resources = []
    for entity in entities:
        entity_value = entity.get('entityValue', '')
        if entity_value:
            resources.append(entity_value)
    
    if resources:
        return ", ".join(resources)
    else:
        return "None specified"

def analyze_event_with_bedrock(bedrock_client, event_data):
    """
    Analyze an AWS Health event using Amazon Bedrock with focus on outage impact
    
    Args:
        bedrock_client: Amazon Bedrock client
        event_data (dict): Event data to analyze
        
    Returns:
        dict: Analyzed event data
    """
    try:
        # Get event details
        event_type = event_data.get('eventTypeCode', event_data.get('event_type', 'Unknown'))
        event_category = event_data.get('eventTypeCategory', event_data.get('event_type_category', 'Unknown'))
        region = event_data.get('region', 'Unknown')
        
        # Format start time if it's a datetime object
        start_time = event_data.get('startTime', event_data.get('start_time', 'Unknown'))
        if hasattr(start_time, 'isoformat'):
            start_time = start_time.isoformat()
        
        # Use description for analysis
        description = event_data.get('description', 'No description available')

        
        # Prepare prompt for Bedrock - ENHANCED FOR OUTAGE ANALYSIS
        print(f"Processing event: {event_type} with category {event_category}")
        print(f"Using description (length: {len(description)}): {description[:100]}...")
        
        prompt = f"""
        You are an AWS expert specializing in outage analysis and business continuity. Your task is to analyze this AWS Health event and determine its potential impact on workload availability, system connectivity, and service outages.
        
        AWS Health Event:
        - Type: {event_type}
        - Category: {event_category}
        - Region: {region}
        - Start Time: {start_time}
        
        Event Description:
        {description}
        
        IMPORTANT ANALYSIS FOCUS:
        1. Will this event cause workload downtime if required actions are not taken?
        2. Will there be any service outages associated with this event?
        3. Will the application/workload experience network integration issues between connecting systems?
        4. What specific AWS services or resources could be impacted?
         
        
        CRITICAL EVENT CRITERIA:
        - Any event that will cause service downtime should be marked as CRITICAL
        - Any event that will cause network integration or SSL issues between systems should be marked as CRITICAL
        - Any event that requires immediate action to prevent outage should be marked as URGENT time sensitivity
        - Events with high impact but no immediate downtime should be marked as HIGH risk level
    
        Please analyze this event and provide the following information in JSON format:
        {{
          "critical": boolean,
          "risk_level": "critical|high|medium|low",
          "account_impact": "critical|high|medium|low",
          "time_sensitivity": "Routine|Urgent|Critical",
          "risk_category": "Availability|Security|Performance|Cost|Compliance",
          "required_actions": "string",
          "impact_analysis": "string",
          "consequences_if_ignored": "string",
          "affected_resources": "string"
        }}
        
        IMPORTANT: In your impact_analysis field, be very specific about:
        1. Potential outages and their estimated duration
        2. Connectivity issues between systems
        3. Whether this will cause downtime if actions are not taken
        
        In your consequences_if_ignored field, clearly state what outages or disruptions will occur if the event is not addressed.

        RISK LEVEL GUIDELINES:
        - CRITICAL: Will cause service outage or severe disruption if not addressed
        - HIGH: Significant impact but not an immediate outage
        - MEDIUM: Moderate impact requiring attention
        - LOW: Minimal impact, routine maintenance
        """
        
        # Determine which model we're using and format accordingly
        model_id = os.environ.get('BEDROCK_MODEL_ID', 'anthropic.claude-v2')
        max_tokens = int(os.environ.get('BEDROCK_MAX_TOKENS', '4000'))
        temperature = float(os.environ.get('BEDROCK_TEMPERATURE', '0.2'))
        top_p = float(os.environ.get('BEDROCK_TOP_P', '0.9'))
        
        print(f"Sending request to Bedrock model: '{model_id}'")
        
        if "claude-3" in model_id.lower():
            # Claude 3 models use the messages format
            payload = {
                "modelId": model_id,
                "contentType": "application/json",
                "accept": "application/json",
                "body": json.dumps({
                    "anthropic_version": "bedrock-2023-05-31",
                    "max_tokens": max_tokens,
                    "temperature": temperature,
                    "top_p": top_p,
                    "messages": [
                        {
                            "role": "user",
                            "content": prompt
                        }
                    ]
                })
            }
        else:
            # Claude 2 and other models use the older prompt format
            payload = {
                "modelId": model_id,
                "contentType": "application/json",
                "accept": "application/json",
                "body": json.dumps({
                    "prompt": f"\n\nHuman: {prompt}\n\nAssistant:",
                    "max_tokens_to_sample": max_tokens,
                    "temperature": temperature,
                    "top_p": top_p
                })
            }
        
        # Call Bedrock
        try:
            response = bedrock_client.invoke_model(**payload)
            response_body = json.loads(response.get('body').read())
            
            # Extract response based on model
            if "claude-3" in model_id.lower():
                response_text = response_body.get('content', [{}])[0].get('text', '')
            else:
                response_text = response_body.get('completion', '')
            
            # Store the full analysis text as a string
            event_data['analysis_text'] = response_text
            
            # Try to extract JSON from the response
            json_match = re.search(r'```json\s*(.*?)\s*```', response_text, re.DOTALL)
            if json_match:
                json_str = json_match.group(1)
            else:
                json_match = re.search(r'({.*})', response_text, re.DOTALL)
                if json_match:
                    json_str = json_match.group(1)
                else:
                    json_str = response_text
            
            # Parse the JSON
            try:
                analysis = json.loads(json_str)
                # NEW CODE: Normalize risk level to ensure consistency
                if 'risk_level' in analysis:
                    risk_level = analysis['risk_level'].strip().upper()
                    
                    # Ensure "critical" is properly recognized and distinguished from "high"
                    if risk_level in ['CRITICAL', 'SEVERE']:
                        analysis['risk_level'] = 'CRITICAL'
                        # Make sure critical boolean flag is consistent
                        analysis['critical'] = True
                    elif risk_level == 'HIGH':
                        analysis['risk_level'] = 'HIGH'
                    elif risk_level in ['MEDIUM', 'MODERATE']:
                        analysis['risk_level'] = 'MEDIUM'
                    elif risk_level == 'LOW':
                        analysis['risk_level'] = 'LOW'
                    
                    # If critical flag is True but risk_level isn't CRITICAL, fix it
                    if analysis.get('critical', False) and analysis['risk_level'] != 'CRITICAL':
                        analysis['risk_level'] = 'CRITICAL'
                
                # Update event data with analysis
                event_data.update(analysis)
                
                return event_data
            except json.JSONDecodeError:
                print(f"Failed to parse JSON from response: {response_text[:200]}...")
                # Provide default values if parsing fails
                event_data.update({
                    'critical': False,
                    'risk_level': 'low',
                    'account_impact': 'low',
                    'time_sensitivity': 'Routine',
                    'risk_category': 'Unknown',
                    'required_actions': 'Review event details manually',
                    'impact_analysis': 'Unable to automatically analyze this event',
                    'consequences_if_ignored': 'Unknown',
                    'affected_resources': 'Unknown'
                })
                return event_data
                
        except Exception as e:
            print(f"Error in Bedrock analysis: {str(e)}")
            traceback.print_exc()
            
            # Provide default values if Bedrock analysis fails
            event_data.update({
                'critical': False,
                'risk_level': 'low',
                'account_impact': 'low',
                'time_sensitivity': 'Routine',
                'risk_category': 'Unknown',
                'required_actions': 'Review event details manually',
                'impact_analysis': 'Unable to automatically analyze this event',
                'consequences_if_ignored': 'Unknown',
                'affected_resources': 'Unknown',
                'analysis_text': f"Error during analysis: {str(e)}"
            })
            return event_data
    
    except Exception as e:
        print(f"Unexpected error in analyze_event_with_bedrock: {str(e)}")
        traceback.print_exc()
        
        # Provide default values if function fails
        event_data.update({
            'critical': False,
            'risk_level': 'low',
            'account_impact': 'low',
            'time_sensitivity': 'Routine',
            'risk_category': 'Unknown',
            'required_actions': 'Review event details manually',
            'impact_analysis': 'Unable to automatically analyze this event',
            'consequences_if_ignored': 'Unknown',
            'affected_resources': 'Unknown',
            'analysis_text': f"Error during analysis: {str(e)}"
        })
        return event_data



def categorize_analysis(analysis_text):
    """
    Extract structured data from Bedrock analysis text
    
    Args:
        analysis_text: Analysis text from Bedrock (string or dict)
        
    Returns:
        dict: Structured data extracted from analysis
    """
    categories = {
        'critical': False,
        'risk_level': 'low',
        'impact_analysis': '',
        'required_actions': '',
        'time_sensitivity': 'Routine',
        'risk_category': 'Unknown',
        'consequences_if_ignored': '',
        'event_category': 'Low'
    }
    
    try:
        # If analysis_text is already a dictionary, use it directly
        if isinstance(analysis_text, dict):
            # Update our categories with values from the dictionary
            for key in categories.keys():
                if key in analysis_text:
                    categories[key] = analysis_text[key]
            
            # Also check for affected_resources
            if 'affected_resources' in analysis_text:
                categories['affected_resources'] = analysis_text['affected_resources']
                
            return categories
            
        # If analysis_text is not a string, convert it to string
        if not isinstance(analysis_text, str):
            analysis_text = str(analysis_text)
            
        # Try to parse as JSON first
        try:
            json_data = json.loads(analysis_text)
            # If successful, update our categories with values from the JSON
            for key in categories.keys():
                if key in json_data:
                    categories[key] = json_data[key]
            
            # Also check for affected_resources
            if 'affected_resources' in json_data:
                categories['affected_resources'] = json_data['affected_resources']
                
            return categories
        except json.JSONDecodeError:
            # Not valid JSON, continue with regex parsing
            pass
            
        # Extract critical status
        critical_match = re.search(r'CRITICAL:\s*(?:\[)?([Yy]es|[Nn]o)(?:\])?', analysis_text)
        if critical_match:
            categories['critical'] = critical_match.group(1).lower() == 'yes'
        
        # Extract risk level
        risk_match = re.search(r'RISK LEVEL:\s*(?:\[)?([Hh]igh|[Mm]edium|[Ll]ow)(?:\])?', analysis_text)
        if risk_match:
            categories['risk_level'] = risk_match.group(1).lower()
        
        # Extract account impact
        impact_match = re.search(r'ACCOUNT IMPACT:\s*(?:\[)?([Hh]igh|[Mm]edium|[Ll]ow)(?:\])?', analysis_text)
        if impact_match:
            categories['account_impact'] = impact_match.group(1).lower()
        
        # Extract impact analysis
        impact_analysis_match = re.search(r'IMPACT ANALYSIS:(.*?)(?:REQUIRED ACTIONS:|$)', analysis_text, re.DOTALL)
        if impact_analysis_match:
            categories['impact_analysis'] = impact_analysis_match.group(1).strip()
        
        # Extract required actions
        required_actions_match = re.search(r'REQUIRED ACTIONS:(.*?)(?:TIME SENSITIVITY:|$)', analysis_text, re.DOTALL)
        if required_actions_match:
            categories['required_actions'] = required_actions_match.group(1).strip()
        
        # Extract time sensitivity
        time_sensitivity_match = re.search(r'TIME SENSITIVITY:\s*([Ii]mmediate|[Uu]rgent|[Ss]oon|[Rr]outine)', analysis_text)
        if time_sensitivity_match:
            categories['time_sensitivity'] = time_sensitivity_match.group(1).capitalize()
        
        # Extract risk category
        risk_category_match = re.search(r'RISK CATEGORY:\s*([Tt]echnical|[Oo]perational|[Ss]ecurity|[Cc]ompliance|[Cc]ost|[Aa]vailability)', analysis_text)
        if risk_category_match:
            categories['risk_category'] = risk_category_match.group(1).capitalize()
        
        # Extract consequences if ignored
        consequences_match = re.search(r'CONSEQUENCES IF IGNORED:(.*?)(?:$)', analysis_text, re.DOTALL)
        if consequences_match:
            categories['consequences_if_ignored'] = consequences_match.group(1).strip()
        
        # Extract affected resources
        affected_match = re.search(r'AFFECTED RESOURCES:(.*?)(?:$)', analysis_text, re.DOTALL)
        if affected_match:
            categories['affected_resources'] = affected_match.group(1).strip()
            
    except Exception as e:
        print(f"Error categorizing analysis: {str(e)}")
    
    return categories


def create_excel_report_improved(events_analysis):
    """
    Create an improved Excel report with detailed event analysis
    
    Args:
        events_analysis (list): List of analyzed events data
        
    Returns:
        BytesIO: Excel file as bytes
    """
    # Create workbook and sheets
    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    events_sheet = wb.create_sheet(title="All Events")
    
    # Add summary data
    summary_sheet['A1'] = "AWS Health Events Analysis Summary"
    summary_sheet['A1'].font = Font(size=16, bold=True)
    summary_sheet.merge_cells('A1:C1')
    
    # Add event counts by category
    event_categories = defaultdict(int)
    for event in events_analysis:
        event_categories[event.get('event_type_category', 'unknown')] += 1
        if event.get('critical', False):
            event_categories['critical'] += 1
        
        risk_level = event.get('risk_level', 'low').lower()
        if risk_level == 'high':
            event_categories['high_risk'] += 1
        elif risk_level == 'medium':
            event_categories['medium_risk'] += 1
        elif risk_level == 'low':
            event_categories['low_risk'] += 1
            
        impact = event.get('account_impact', 'low').lower()
        if impact == 'high':
            event_categories['high_impact'] += 1
        elif impact == 'medium':
            event_categories['medium_impact'] += 1
        elif impact == 'low':
            event_categories['low_impact'] += 1
    
    # Add summary counts
    summary_sheet['A3'] = "Event Categories"
    summary_sheet['A3'].font = Font(bold=True)
    
    row = 4
    for category, count in event_categories.items():
        summary_sheet[f'A{row}'] = category.replace('_', ' ').title()
        summary_sheet[f'B{row}'] = count
        row += 1
    
    # Add headers to events sheet
    headers = [
        "Event ARN",  # Added Event ARN as first column
        "Event Type", 
        "Region", 
        "Start Time", 
        "Last Update", 
        "Category",
        "Description",  # Added Description column
        "Critical", 
        "Risk Level", 
        "Account ID", 
        "Time Sensitivity", 
        "Risk Category", 
        "Required Actions", 
        "Impact Analysis", 
        "Consequences If Ignored", 
        "Affected Resources"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = events_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Add event data
    for row_num, event in enumerate(events_analysis, 2):
        event_arn = event.get('eventArn', event.get('arn', 'N/A'))
        events_sheet.cell(row=row_num, column=1).value = event_arn # Add Event ARN
        events_sheet.cell(row=row_num, column=2).value = event.get('event_type', 'N/A')
        events_sheet.cell(row=row_num, column=3).value = event.get('region', 'N/A')
        events_sheet.cell(row=row_num, column=4).value = event.get('start_time', 'N/A')
        events_sheet.cell(row=row_num, column=5).value = event.get('last_update_time', 'N/A')
        events_sheet.cell(row=row_num, column=6).value = event.get('event_type_category', 'N/A')
        
        # Add description with text wrapping
        description_cell = events_sheet.cell(row=row_num, column=7)
        description_cell.value = event.get('description', 'N/A')
        description_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        events_sheet.cell(row=row_num, column=8).value = "Yes" if event.get('critical', False) else "No"
        events_sheet.cell(row=row_num, column=9).value = event.get('risk_level', 'low').upper()
        events_sheet.cell(row=row_num, column=10).value = event.get('accountId', 'N/A')
        events_sheet.cell(row=row_num, column=11).value = event.get('time_sensitivity', 'Routine')
        events_sheet.cell(row=row_num, column=12).value = event.get('risk_category', 'Unknown')
        events_sheet.cell(row=row_num, column=13).value = event.get('required_actions', '')
        events_sheet.cell(row=row_num, column=14).value = event.get('impact_analysis', '')
        events_sheet.cell(row=row_num, column=15).value = event.get('consequences_if_ignored', '')
        events_sheet.cell(row=row_num, column=16).value = event.get('affected_resources', 'None')
        
        # Color coding based on risk level
        risk_level = event.get('risk_level', 'low').lower()
        is_critical = event.get('critical', False)
        
        if is_critical:
            for col_num in range(1, 17):
                events_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )
        elif risk_level == 'high':
            for col_num in range(1, 17):
                events_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                )
        elif risk_level == 'medium':
            for col_num in range(1, 17):
                events_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="E6F2FF", end_color="E6F2FF", fill_type="solid"
                )
    
    # Auto-adjust column widths
    for col in events_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 50)  # Cap at 50 characters
                except:
                    pass
        adjusted_width = max_length + 2
        events_sheet.column_dimensions[column].width = adjusted_width
    
    # Set specific width for description column
    events_sheet.column_dimensions['G'].width = 60  # Description column
    
    # Create critical events sheet
    critical_sheet = wb.create_sheet(title="Critical Events")
    
    # Add headers (same as events sheet)
    for col_num, header in enumerate(headers, 1):
        cell = critical_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Add critical event data
    critical_events = [event for event in events_analysis if event.get('critical', False)]
    if not critical_events:
        critical_sheet.cell(row=2, column=1).value = "No critical events found"
        critical_sheet.merge_cells('A2:P2')  # Updated to include new columns
        critical_sheet.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    else:
        for row_num, event in enumerate(critical_events, 2):
            event_arn = event.get('eventArn', event.get('arn', 'N/A'))
            critical_sheet.cell(row=row_num, column=1).value = event_arn
            critical_sheet.cell(row=row_num, column=2).value = event.get('event_type', 'N/A')
            critical_sheet.cell(row=row_num, column=3).value = event.get('region', 'N/A')
            critical_sheet.cell(row=row_num, column=4).value = event.get('start_time', 'N/A')
            critical_sheet.cell(row=row_num, column=5).value = event.get('last_update_time', 'N/A')
            critical_sheet.cell(row=row_num, column=6).value = event.get('event_type_category', 'N/A')
            
            # Add description with text wrapping
            description_cell = critical_sheet.cell(row=row_num, column=7)
            description_cell.value = event.get('description', 'N/A')
            description_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            critical_sheet.cell(row=row_num, column=8).value = "Yes"
            critical_sheet.cell(row=row_num, column=9).value = event.get('risk_level', 'low').upper()
            critical_sheet.cell(row=row_num, column=10).value = event.get('accountId', 'N/A')
            critical_sheet.cell(row=row_num, column=11).value = event.get('time_sensitivity', 'Routine')
            critical_sheet.cell(row=row_num, column=12).value = event.get('risk_category', 'Unknown')
            critical_sheet.cell(row=row_num, column=13).value = event.get('required_actions', '')
            critical_sheet.cell(row=row_num, column=14).value = event.get('impact_analysis', '')
            critical_sheet.cell(row=row_num, column=15).value = event.get('consequences_if_ignored', '')
            critical_sheet.cell(row=row_num, column=16).value = event.get('affected_resources', 'None')
            
            # Apply critical highlighting
            for col_num in range(1, 17):
                critical_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )
    
    # Auto-adjust column widths for critical sheet
    for col in critical_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 50)  # Cap at 50 characters
                except:
                    pass
        adjusted_width = max_length + 2
        critical_sheet.column_dimensions[column].width = adjusted_width
    
    # Set specific width for description column in critical sheet
    critical_sheet.column_dimensions['G'].width = 60  # Description column
    
     # Create risk analysis sheet with full Bedrock analysis
    analysis_sheet = wb.create_sheet(title="Risk Analysis")
    
    # Add headers
    analysis_headers = ["Event Type", "Region", "Risk Level", "Full Analysis"]
    
    for col_num, header in enumerate(analysis_headers, 1):
        cell = analysis_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Add analysis data
    for row_num, event in enumerate(events_analysis, 2):
        # Get event type from either format
        event_type = event.get('eventTypeCode', event.get('event_type', 'N/A'))
        analysis_sheet.cell(row=row_num, column=1).value = event_type
        
        # Get region
        analysis_sheet.cell(row=row_num, column=2).value = event.get('region', 'N/A')
        
        # Get risk level
        analysis_sheet.cell(row=row_num, column=3).value = event.get('risk_level', 'low').upper()
        
        # Fix: Ensure analysis_text is a string
        analysis_text = event.get('analysis_text', '')
        
        # Handle different data types
        if isinstance(analysis_text, dict):
            # Custom JSON encoder to handle datetime objects
            def datetime_handler(obj):
                if isinstance(obj, (datetime, date)):
                    return obj.isoformat()
                return str(obj)
            
            try:
                analysis_text = json.dumps(analysis_text, indent=2, default=datetime_handler)
            except Exception as e:
                analysis_text = f"Error serializing analysis: {str(e)}"
        elif not isinstance(analysis_text, str):
            # Convert non-string values to string
            try:
                analysis_text = str(analysis_text)
            except Exception as e:
                analysis_text = f"Error converting to string: {str(e)}"
        
        # Set the cell value
        analysis_sheet.cell(row=row_num, column=4).value = analysis_text
        
        # Color coding based on risk level
        risk_level = event.get('risk_level', 'low').lower()
        is_critical = event.get('critical', False)
        
        if is_critical:
            for col_num in range(1, 5):
                analysis_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )
        elif risk_level == 'high':
            for col_num in range(1, 5):
                analysis_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                )
        elif risk_level == 'medium':
            for col_num in range(1, 5):
                analysis_sheet.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color="E6F2FF", end_color="E6F2FF", fill_type="solid"
                )
    # Set column widths for analysis sheet
    analysis_sheet.column_dimensions['A'].width = 30
    analysis_sheet.column_dimensions['B'].width = 15
    analysis_sheet.column_dimensions['C'].width = 15
    analysis_sheet.column_dimensions['D'].width = 100
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)


    
    return excel_buffer

def generate_summary_html(total_events, event_categories, filtered_events, category_filter, events_analysis):
    """
    Generate HTML summary for email
    
    Args:
        total_events (int): Total number of events
        event_categories (dict): Event categories count
        filtered_events (int): Number of filtered events
        category_filter (list): Categories used for filtering
        events_analysis (list): Analyzed events data
        
    Returns:
        str: HTML content for email
    """
    # Current date for the report
    current_date = datetime.now().strftime('%Y-%m-%d')
    
    # Calculate accurate event counts directly from events_analysis
    critical_count = sum(1 for event in events_analysis if event.get('critical', False))
    high_risk_count = sum(1 for event in events_analysis if event.get('risk_level', '').lower() == 'high')
    medium_risk_count = sum(1 for event in events_analysis if event.get('risk_level', '').lower() == 'medium')
    low_risk_count = sum(1 for event in events_analysis if event.get('risk_level', '').lower() == 'low')

    # Define CSS for better table formatting
    table_css = """
    <style>
        .health-events-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 5px;
        }
        .health-events-table th, .health-events-table td {
            border: 1px solid #ddd;
            padding: 6px;
            text-align: left;
            vertical-align: top;
            word-wrap: break-word;
            font-size: 5px;
        }
        .health-events-table th {
            background-color: #f2f2f2;
        }
        /* Column width constraints */
        .col-arn {
            width: 15%;
        }
        .col-region {
            width: 8%;
        }
        .col-start-time {
            width: 12%;
        }
        .col-risk {
            width: 8%;
        }
        .col-accountid {
            width: 12%;
        }
    </style>
    """
    
    # Start building HTML content
    html_content = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; }}
            .header {{ background-color: #232F3E; color: white; padding: 20px; }}
            .content {{ padding: 20px; }}
            table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left;"font-size:5px;" }}
            th {{ background-color: #f2f2f2; }}
            .critical {{ background-color: #ffcccc; }}
            .high {{ background-color: #fff2cc; }}
            .medium {{ background-color: #e6f2ff; }}
            .summary {{ margin-top: 20px; margin-bottom: 20px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>AWS Health Events Analysis Report</h1>
            <p>Date: {current_date}</p>
        </div>
        <div class="content">
            <div class="summary">
                <h2>Summary</h2>
                <p>Total AWS Health events analyzed: {len(events_analysis)} of {total_events} events found</p>
    """
    
    # Add analysis window information
    end_time = datetime.utcnow()
    start_time = end_time - timedelta(days=int(os.environ['ANALYSIS_WINDOW_DAYS']))
    html_content += f"""
                <p>Analysis Window: {start_time.strftime('%Y-%m-%d %H:%M:%S')} UTC to {end_time.strftime('%Y-%m-%d %H:%M:%S')} UTC</p>
    """
    
    # Add filter information if applicable
    if category_filter:
        html_content += f"""
                <p>Events filtered by categories: {', '.join(category_filter)}</p>
                <p>Events excluded by filter: {filtered_events}</p>
        """
    
    # Add event category counts - USING ACCURATE COUNTS
    html_content += """
                <h3>Event Categories</h3>
                <ul>
    """
    
    # Add critical events count if any
    if critical_count > 0:
        html_content += f"""
                <li><strong>Critical Events:</strong> {critical_count}</li>
        """
    
    # Add risk level counts - USING ACCURATE COUNTS
    html_content += f"""
                <li><strong>High Risk Events:</strong> {high_risk_count}</li>
                <li><strong>Medium Risk Events:</strong> {medium_risk_count}</li>
                <li><strong>Low Risk Events:</strong> {low_risk_count}</li>
            </ul>
            </div>
    """
    
    # Add critical events table if any exist
    critical_events = [event for event in events_analysis if event.get('critical', False)]
    if critical_events:
        html_content += """
            <h2>Critical Events</h2>
            <table class="health-events-table">
                <tr>
                    <th style="font-size:12px;">Event ARN</th>
                    <th style="font-size:12px;">Region</th>
                    <th style="font-size:12px;">Start Time</th>
                    <th style="font-size:12px;">Risk Level</th>
                    <th style="font-size:12px;">Account ID</th>
                </tr>
        """
        
        for event in critical_events:
            # Get event ARN, preferring eventArn if available, falling back to arn
            event_arn = event.get('eventArn', event.get('arn', 'N/A'))
            
            html_content += f"""
                <tr class="critical">
                    <td style="font-size:12px;">{event_arn}</td>
                    <td style="font-size:12px;">{event.get('region', 'N/A')}</td>
                    <td style="font-size:12px;">{event.get('start_time', 'N/A')}</td>
                    <td style="font-size:12px;">{event.get('risk_level', 'N/A').upper()}</td>
                    <td style="font-size:12px;">{event.get('accountId', 'N/A')}</td>
                </tr>
            """
        
        html_content += """
            </table>
        """
    
    # Add high risk events table
    high_risk_events = [event for event in events_analysis if event.get('risk_level', '').lower() == 'high']
    if high_risk_events:
        html_content += """
            <h2>High Risk Events</h2>
            <table class="health-events-table">
                <tr>
                    <th style="font-size:12px;">Event ARN</th>
                    <th style="font-size:12px;">Region</th>
                    <th style="font-size:12px;">Start Time</th>
                    <th style="font-size:12px;">Risk Level</th>
                    <th style="font-size:12px;">Account ID</th>
                </tr>
        """
        
        for event in high_risk_events:
            # Get event ARN, preferring eventArn if available, falling back to arn
            event_arn = event.get('eventArn', event.get('arn', 'N/A'))
            
            html_content += f"""
                <tr class="high">
                    <td style="font-size:12px;">{event_arn}</td>
                    <td style="font-size:12px;">{event.get('region', 'N/A')}</td>
                    <td style="font-size:12px;">{event.get('start_time', 'N/A')}</td>
                    <td style="font-size:12px;">{event.get('risk_level', 'N/A').upper()}</td>
                    <td style="font-size:12px;">{event.get('accountId', 'N/A')}</td>
                </tr>
            """
        
        html_content += """
            </table>
        """
    
    # Add footer with attachment information
    html_content += """
            <div class="summary">
                <h2>Full Report</h2>
                <p>Please see the attached Excel file for complete details on all events.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return html_content


def send_ses_email_with_attachment(html_content, excel_buffer, total_events, event_categories,events_analysis):
    """
    Send email with Excel attachment using Amazon SES
    
    Args:
        html_content (str): HTML content for email body
        excel_buffer (BytesIO): Excel file as bytes
        total_events (int): Total number of events
        event_categories (dict): Event categories count
        
    Returns:
        None
    """
    try:
        # Get email configuration from environment variables
        sender = os.environ['SENDER_EMAIL']
        recipients_str = os.environ['RECIPIENT_EMAILS']
        recipients = [email.strip() for email in recipients_str.split(',')]
        
        # Create email subject with counts
        critical_count = event_categories.get('critical', 0)
        #high_risk_count = event_categories.get('high_risk', 0)
        #high_risk_count = event_categories.get(('risk_level', '').lower() == 'high'),0)
        high_risk_count = sum(1 for event in events_analysis if event.get('risk_level', '').lower() == 'high')
        
        if critical_count > 0:
            subject = f"{customer_name} [CRITICAL] AWS Health Events Analysis - {critical_count} Critical, {high_risk_count} High Risk Events"
        elif high_risk_count > 0:
            subject = f"{customer_name} [HIGH RISK] AWS Health Events Analysis - {high_risk_count} High Risk Events"
        else:
            subject = f"{customer_name} AWS Health Events Analysis - {total_events} Events"
        
        # Generate Excel filename
                # Generate Excel filename
        current_date = datetime.now().strftime('%Y-%m-%d')
        current_time = datetime.now().strftime('%H-%M-%S')
        excel_filename = EXCEL_FILENAME_TEMPLATE.format(date=current_date, time=current_time)
        
        # Create SES client
        ses_client = boto3.client('ses')
        
        # Create message container
        message = {
            'Subject': {
                'Data': subject
            },
            'Body': {
                'Html': {
                    'Data': html_content
                }
            }
        }
        
        # Create raw email message with attachment
        msg_raw = {
            'Source': sender,
            'Destinations': recipients,
            'RawMessage': {
                'Data': create_raw_email_with_attachment(
                    sender=sender,
                    recipients=recipients,
                    subject=subject,
                    html_body=html_content,
                    attachment_data=excel_buffer.getvalue(),
                    attachment_name=excel_filename
                )
            }
        }
        
        # Send email
        response = ses_client.send_raw_email(**msg_raw)
        print(f"Email sent successfully. Message ID: {response['MessageId']}")

        
    except Exception as e:
        print(f"Error sending email: {str(e)}")
        traceback.print_exc()
    
    try:
        # Check if S3 bucket name is configured
        if not S3_BUCKET_NAME:
            return False, "S3_BUCKET_NAME environment variable not configured"
            
        # Create S3 client
        s3_client = boto3.client('s3')
        
        # Generate S3 key with prefix if provided
        s3_key = f"{S3_KEY_PREFIX.rstrip('/')}/{excel_filename}" if S3_KEY_PREFIX else file_name
        
        # Reset buffer position to the beginning
        excel_buffer.seek(0)
        
        # Upload buffer to S3 using put_object
        s3_client.put_object(
            Bucket=S3_BUCKET_NAME,
            Key=s3_key,
            Body=excel_buffer.getvalue(),
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Generate S3 URL for the uploaded file
        s3_url = f"s3://{S3_BUCKET_NAME}/{s3_key}"
        
        print(f"Successfully uploaded file to {s3_url}")
             
    except Exception as e:
        error_message = f"Error uploading file to S3: {str(e)}"
        print(error_message)
        return False, error_message

def create_raw_email_with_attachment(sender, recipients, subject, html_body, attachment_data, attachment_name):
    """
    Create raw email with attachment
    
    Args:
        sender (str): Sender email
        recipients (list): List of recipient emails
        subject (str): Email subject
        html_body (str): HTML email body
        attachment_data (bytes): Attachment data
        attachment_name (str): Attachment filename
        
    Returns:
        bytes: Raw email message
    """
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication
    
    # Create message container
    msg = MIMEMultipart('mixed')
    msg['Subject'] = subject
    msg['From'] = sender
    msg['To'] = ', '.join(recipients)
    
    # Create HTML part
    msg_body = MIMEMultipart('alternative')
    html_part = MIMEText(html_body, 'html')
    msg_body.attach(html_part)
    msg.attach(msg_body)
    
    # Create attachment part
    att = MIMEApplication(attachment_data)
    att.add_header('Content-Disposition', 'attachment', filename=attachment_name)
    msg.attach(att)
    
    # Convert to string and return
    return msg.as_string().encode('utf-8')

def add_cloudwatch_metrics(event_categories, analyzed_count, total_count, filtered_count):
    """
    Add CloudWatch metrics for monitoring
    
    Args:
        event_categories (dict): Event categories count
        analyzed_count (int): Number of analyzed events
        total_count (int): Total number of events
        filtered_count (int): Number of filtered events
        
    Returns:
        None
    """
    try:
        # Create CloudWatch client
        cloudwatch = boto3.client('cloudwatch')
        
        # Create metrics data
        metrics_data = [
            {
                'MetricName': 'AnalyzedEvents',
                'Value': analyzed_count,
                'Unit': 'Count',
                'Dimensions': [
                    {
                        'Name': 'Function',
                        'Value': 'HealthEventsAnalysis'
                    }
                ]
            },
            {
                'MetricName': 'TotalEvents',
                'Value': total_count,
                'Unit': 'Count',
                'Dimensions': [
                    {
                        'Name': 'Function',
                        'Value': 'HealthEventsAnalysis'
                    }
                ]
            },
            {
                'MetricName': 'FilteredEvents',
                'Value': filtered_count,
                'Unit': 'Count',
                'Dimensions': [
                    {
                        'Name': 'Function',
                        'Value': 'HealthEventsAnalysis'
                    }
                ]
            },
            {
                'MetricName': 'CriticalEvents',
                'Value': event_categories.get('critical', 0),
                'Unit': 'Count',
                'Dimensions': [
                    {
                        'Name': 'Function',
                        'Value': 'HealthEventsAnalysis'
                    }
                ]
            },
            {
                'MetricName': 'HighRiskEvents',
                'Value': event_categories.get('high_risk', 0),
                'Unit': 'Count',
                'Dimensions': [
                    {
                        'Name': 'Function',
                        'Value': 'HealthEventsAnalysis'
                    }
                ]
            }
        ]
        
        # Put metrics data
        cloudwatch.put_metric_data(
            Namespace='AWS/HealthEventsAnalysis',
            MetricData=metrics_data
        )
        
        print("CloudWatch metrics published successfully")
        
    except Exception as e:
        print(f"Error publishing CloudWatch metrics: {str(e)}")
        # Don't raise the exception - metrics are non-critical

def fetch_health_event_details1(event_arn, account_id=None):
    """
    Fetch detailed event information from AWS Health API for any account in the organization
    
    Args:
        event_arn (str): ARN of the health event
        account_id (str, optional): AWS account ID that owns the event
        
    Returns:
        dict: Event details including affected resources
    """
    try:
        health_client = boto3.client('health', region_name='us-east-1')
        
        # First try organization API (works for both current and linked accounts)
        try:
            # Prepare request for organization event details
            org_filter = {
                'eventArn': event_arn
            }
            
            # Add account ID if provided
            if account_id:
                org_filter['awsAccountId'] = account_id
            
            # Get event details using organization API
            org_event_details = health_client.describe_event_details_for_organization(
                organizationEventDetailFilters=[org_filter]
            )
            
            # Get affected entities using organization API
            org_affected_entities = health_client.describe_affected_entities_for_organization(
                organizationEntityFilters=[
                    {
                        'eventArn': event_arn,
                        'awsAccountId': account_id if account_id else get_account_id_from_event(event_arn)
                    }
                ]
            )
            
            # Check if we got successful results
            if org_event_details.get('successfulSet') and len(org_event_details['successfulSet']) > 0:
                return {
                    'details': org_event_details['successfulSet'][0],
                    'entities': org_affected_entities.get('entities', [])
                }
            
            # If we got here, organization API didn't return results
            print(f"Organization API didn't return results for event {event_arn}")
            
        except Exception as org_error:
            print(f"Error using organization API for event {event_arn}: {str(org_error)}")
        
        # Fall back to account-specific API (only works for current account)
        print(f"Falling back to account-specific API for event {event_arn}")
        
        event_details = health_client.describe_event_details(
            eventArns=[event_arn]
        )
        
        affected_entities = health_client.describe_affected_entities(
            filter={
                'eventArns': [event_arn]
            }
        )
        
        return {
            'details': event_details.get('successfulSet', [{}])[0] if event_details.get('successfulSet') else {},
            'entities': affected_entities.get('entities', [])
        }
        
    except Exception as e:
        print(f"Error fetching Health API data: {str(e)}")
        return {'details': {}, 'entities': []}

def upload_file_to_s3(file_path, file_name):
    """
    Upload a file to S3 bucket using environment variables for configuration
    
    Args:
        file_path: Local path to the file
        file_name: Name to use for the file in S3
        
    Returns:
        tuple: (success boolean, S3 URL or error message)
    """
    try:
        # Check if S3 bucket name is configured
        if not S3_BUCKET_NAME:
            return False, "S3_BUCKET_NAME environment variable not configured"
            
        # Create S3 client
        s3_client = boto3.client('s3')
        
        # Generate S3 key with prefix if provided
        s3_key = f"{S3_KEY_PREFIX.rstrip('/')}/{file_name}" if S3_KEY_PREFIX else file_name
        
        # Upload file to S3
        s3_client.upload_file(file_path, S3_BUCKET_NAME, s3_key)
        
        # Generate S3 URL for the uploaded file
        s3_url = f"s3://{S3_BUCKET_NAME}/{s3_key}"
        
        print(f"Successfully uploaded file to {s3_url}")
        return True, s3_url
        
    except Exception as e:
        error_message = f"Error uploading file to S3: {str(e)}"
        print(error_message)
        return False, error_message







                
